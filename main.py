#main.py
from fastapi import FastAPI , Request, status, Header, HTTPException, Request
from fastapi.responses import JSONResponse
import requests
import json
import  os
from typing import Annotated
from datetime import datetime
from pydantic import BaseModel
from typing import Optional
from dotenv import load_dotenv
import openpyxl 
import email 
import imaplib
import smtplib
from email.message import EmailMessage
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.message import EmailMessage
import pandas as pd

app = FastAPI()



# Modelo para los datos que recibiremos en el POST
class Item(BaseModel):
    nombre: str
    descripcion: Optional[str] = None
    precio: float

#Veririfica que existan el folder para guardar la info recibida
folder_name = "wms_payloads"
#Comprobar si no existe la carpeta , crearla
if not os.path.exists(folder_name):
    os.makedirs(folder_name)

#Veririfica que existan el folder para guardar la info recibida
folder_excel = "excel"
#Comprobar si no existe la carpeta , crearla
if not os.path.exists(folder_excel):
    os.makedirs(folder_excel)

#Abrir el archivo de configuración para urls y credenciales
from pathlib import Path
ruta_base = Path(__file__).parent 
archivo_ruta = ruta_base / "config.json"
with open(archivo_ruta, 'r') as f:
    data = json.load(f) # Lee y convierte a diccionario en un solo paso
    url_token_tms   = data["items"][0]["url_token_tms"]
    c_id            = data["items"][0]["c_id"]
    c_sec           = data["items"][0]["c_sec"]
    scope           = data["items"][0]["scope"]

#Arma la url del TMS BY token
url = url_token_tms

headers = {
    'Content-Type': 'application/x-www-form-urlencoded'
}

estatus = "Recibido"

@app.post("/test/wms/rc", status_code=status.HTTP_202_ACCEPTED)
async def get_json_raw(request: Request,x_token_key: str = Header(...)):
   
    
    """
    Endpoint POST que requiere 'x-token-key' en el encabezado.
    """
    load_dotenv()
    SECRET_KEY = os.getenv("SECRET_KEY")
    # Validar el token
    if x_token_key != SECRET_KEY:
        raise HTTPException(status_code=403, detail="Token no válido o inexistente")
    
   

    # 1. Leer el stream de bytes
    raw_body = await request.body()
    
    data = json.loads(raw_body)
    trknum =  data["MASTER_RCPT_COMPLETE_OUB_IFD"]["RCV_TRLR_OUB_IFD"]["MASTER_RCPT_OUB_IFD"]["TRKNUM"]
    trknum_limpio = trknum.replace("/", "_")
    print(trknum_limpio)
    trknum_limpio = trknum.replace(".", "_")
    # 2. Parsear manualmente
    try:
        data      = json.loads(raw_body)
        fecha_str = datetime.now().strftime("%Y%m%d%H%M%S")
        archivo   = folder_name+"/"+trknum_limpio+'_'+fecha_str+".json"
        
        with open(archivo, "w", encoding="utf-8") as f:
             json.dump(data, f, indent=4, ensure_ascii=False)
        
        ###asincrono###   
        
        dataF = [("Cliente", "Número de Entrada", "Proveedor", "No. Factura", "No. De producto", "SKU", "Cajas Solicitadas a Recibir", "Flips Solicitados a Recibir", "Cajas Recibidas", "Flips Recibidos", "Estilo", "Batch", "Pack Code", "Fecha de Caducidad", "Contenedor", "Estado de Calidad", "Referencia", "Linea"),
                ]   
            
        lineas = data["MASTER_RCPT_COMPLETE_OUB_IFD"]["RCV_TRLR_OUB_IFD"]["MASTER_RCPT_OUB_IFD"]["RCPT_INVOICE_OUB_IFD"]["RCPT_INVOICE_LINE_OUB_IFD"]  
        SUPNUM = data["MASTER_RCPT_COMPLETE_OUB_IFD"]["RCV_TRLR_OUB_IFD"]["MASTER_RCPT_OUB_IFD"]["RCPT_INVOICE_OUB_IFD"]["SUPNUM"]

        # 1. Inicializas una variable para rastrear el valor previo
        ultimo_invsln = None
        ultimo_EXPQTY = 0
        EXPQTY        = 0
        CAJASR        = 0      
        qtyFlips      = 0
        CAJASRR       = 0
        EXPQTY2       = 0
        ultimo_qtyFlips = 0
        rcvsts        = "DISPONIBLE"

        CLIENT_ID   = ""
        NUM_FACTURA = "" 
        EAN         = ""
       
        LOTNUM      =""
        SUP_LOTNUM  =""
        PACK_CODE   ="" 
        expire_dte  =""
        PEDIMENTO   =""
        REFERENCIA  =""

        ultimo_REFERENCIA  = ""
        ultimo_PACK_CODE   = ""  
        ultimo_PEDIMENTO   = ""
        ultimo_NUM_FACTURA = ""
        ultimo_EAN         = ""
        ultimo_expire_dte  = ""
        ultimo_CLIENT_ID   = ""
        ultimo_LOTNUM      = ""
        ultimo_SUP_LOTNUM  = ""
        ultimo_rcvsts      = ""


        for linea in lineas:
            # 2. Obtienes el valor actual
            valor_actual = linea.get("INVSLN")
            invsln = linea.get("INVSLN")
            print(f"Valor de INVSLN: {invsln}")           
             
            PRTNUM       = linea.get("PRTNUM")
            if linea.get("INV_ATTR_STR2"):
                PACK_CODE    = linea.get("INV_ATTR_STR2")

            if linea.get("INV_ATTR_STR3"):
                REFERENCIA = linea.get("INV_ATTR_STR3")
            if linea.get("INV_ATTR_STR4"):    
                PEDIMENTO = linea.get("INV_ATTR_STR4")
            if linea.get("INV_ATTR_STR5"):
                NUM_FACTURA  = linea.get("INV_ATTR_STR5")
            if linea.get("INV_ATTR_STR6"):
                EAN  = linea.get("INV_ATTR_STR6")
            if linea.get("INV_ATTR_STR7"):
                expire_dte  = linea.get("INV_ATTR_STR7")
            if linea.get("CLIENT_ID"):
                CLIENT_ID = linea.get("CLIENT_ID")
            if linea.get("LOTNUM"):
                LOTNUM = linea.get("LOTNUM")
            if linea.get("SUP_LOTNUM"):
                SUP_LOTNUM = linea.get("SUP_LOTNUM")

            if linea.get("RCVSTS"):
                RCVSTS = linea.get("RCVSTS")				
                match RCVSTS:                 
                    case 'A':
                        rcvsts ="DISPONIBLE"
                    case 'RESV':
                        rcvsts = "RESERVA"
                    case 'OBSE':
                        rcvsts ="OBSOLETO"
                    case 'EXP':
                        rcvsts ="EXPIRED"
                    case 'RESG':
                        rcvsts ="RESGUARDO"
            
            if linea.get("RCVQTY"):
               qtyFlips = linea.get("RCVQTY")  
             
           
            if linea.get("EXPQTY"):  
               EXPQTY = linea.get("EXPQTY")  
            

            # ¡Aquí detectas el cambio! 
             # 3. Comparas: si es distinto al último guardado (y no es el primero)
            if ultimo_invsln is not None and valor_actual != ultimo_invsln:
                print(f"El x valor cambió de {ultimo_invsln} a {valor_actual}. Ejecutando acción...")                         
                nuevos_datos = [(ultimo_CLIENT_ID, trknum, SUPNUM, ultimo_NUM_FACTURA, ultimo_EAN, ultimo_PRTNUM,CAJASR,ultimo_EXPQTY,CAJASRR,ultimo_qtyFlips,ultimo_LOTNUM,ultimo_SUP_LOTNUM,ultimo_PACK_CODE, ultimo_expire_dte,ultimo_PEDIMENTO,ultimo_rcvsts,ultimo_REFERENCIA,ultimo_invsln  )
                      ]
                dataF.extend(nuevos_datos)	
                ultimo_qtyFlips =0
                ultimo_EXPQTY   =0
              
            
            # 4. Actualizas la variable auxiliar para la siguiente vuelta               
            if linea.get("EXPQTY"):
               ultimo_EXPQTY      += EXPQTY
               load_dotenv()                
               val_env = os.getenv(linea.get("PRTNUM"), "0")
               print(f"val_env {val_env}")
               if val_env and val_env!= 0:
                  CAJASEXP = float(val_env)
               else: 
                  CAJASEXP = 50           
               
               CAJASEXP = int(CAJASEXP)
               print(f"CAJASEXP {CAJASEXP}")       
               
               if ultimo_EXPQTY !=0 and CAJASEXP !=0:
                  CAJASR             =ultimo_EXPQTY/CAJASEXP
               else:
                  CAJASR  =1       
               
            if linea.get("RCVQTY"):
               valor =linea.get("RCVQTY")
               print(f"rcvqty- {valor }")
               ultimo_qtyFlips += linea.get("RCVQTY")
               print(f"ultimo_qtyFlips- {ultimo_qtyFlips }")
               load_dotenv()                
               val_env = os.getenv(linea.get("PRTNUM"), "0")
               if val_env:
                  CAJASEXP = float(val_env)
               else: 
                  CAJASEXP = 50       
               CAJASEXP = int(CAJASEXP)
               print(f"CAJASEXP {CAJASEXP}")                

               if ultimo_qtyFlips !=0 and CAJASEXP !=0:
                  CAJASRR  = (ultimo_qtyFlips/CAJASEXP) 
               else:   
                   CAJASR  =1     
            ultimo_invsln      = valor_actual
            ultimo_PRTNUM      = PRTNUM
            ultimo_PACK_CODE   = PACK_CODE 
            ultimo_PEDIMENTO   = PEDIMENTO
            ultimo_NUM_FACTURA = NUM_FACTURA
            ultimo_EAN         = EAN
            ultimo_expire_dte  = expire_dte
            ultimo_CLIENT_ID   = CLIENT_ID
            ultimo_LOTNUM      = LOTNUM
            ultimo_SUP_LOTNUM  = SUP_LOTNUM
            ultimo_rcvsts      = rcvsts
            ultimo_REFERENCIA  = REFERENCIA

        nuevos_datos = [(ultimo_CLIENT_ID, trknum, SUPNUM, ultimo_NUM_FACTURA, ultimo_EAN, ultimo_PRTNUM,CAJASR,ultimo_EXPQTY,CAJASRR,ultimo_qtyFlips,ultimo_LOTNUM,ultimo_SUP_LOTNUM,ultimo_PACK_CODE, ultimo_expire_dte,ultimo_PEDIMENTO,ultimo_rcvsts,ultimo_REFERENCIA,ultimo_invsln  )
                   ]
        dataF.extend(nuevos_datos)	
               
        #crear el excel        
        excel_book = openpyxl.Workbook()        
        sheet      = excel_book.active
        sheet.title = trknum_limpio
        
        for index_fila, row in enumerate(dataF):
           for index_col, valor in enumerate(row):
            # index_col + 1 porque Excel empieza en columna 1
              sheet.cell(row=index_fila + 1, column=index_col + 1, value=valor)
         
            
          
        header_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
        header_font = Font(bold=True)
        
        #Aplicar a la primera fila (ajusta el rango según tus columnas, ej: A a N)
        for cell in sheet[1]: 
           cell.fill = header_fill
           cell.font = header_font
        
        #Definir cómo será la línea del borde (delgada y negra)
        thin_border = Border(
                             left=Side(style='thin'),  
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin')
                             )
        #Recorrer todas las celdas que tienen datos y aplicar el borde
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
          for cell in row:
             cell.border = thin_border



        #Recorrer cada columna de la hoja
        for col in sheet.columns:
             max_length = 0
             column_letter = get_column_letter(col[0].column) # Obtiene 'A', 'B', etc.

             #Medir el texto más largo en esa columna
             for cell in col:
                try:
                   if cell.value:
                    # Comparamos la longitud actual con la del valor de la celda
                     if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
    
             #Ajustar el ancho (agregamos 2 unidades de margen)
             adjusted_width = (max_length + 2)
             sheet.column_dimensions[column_letter].width = adjusted_width


        excel_book.save(f"{folder_excel}\{trknum_limpio}.xlsx") 
       
        imap_server =os.getenv("imap_server")
        impap_port  =os.getenv("impap_port")
        username= os.getenv("username2")
        password= os.getenv("password2")
        print(imap_server)
        print(impap_port)
        print(username)
        print(password)

        #Conexion al servidor de correo
        mail = imaplib.IMAP4_SSL(imap_server)
        mail.login(username,password)

        
        original_subject = "-Confirmación de Recibo Envío entrante:"+trknum
        #original_to    = "garcia.miguel@dickalogistics.com.mx"
        #original_cc      ="garcia.miguel@dickalogistics.com.mx"
        original_to    = "Olga.Bohorquez@jti.com"
        original_cc      = "garcia.miguel@dickalogistics.com.mx,Liliana.Cervantes@jti.com,Arturo.Olivares@jti.com,Uriel.Sanchez@jti.com,c06.jefeoperaciones@dickalogistics.com.mx,Sup.tepotzotlan@dickalogistics.com.mx,sup.tepotzotlan@dickalogistics.com.mx,hernandez.guadalupe@dickalogistics.com.mx"
        




        
        #Crear el nuevo mensaje de respuesta
        reply = EmailMessage()
        reply['Subject'] = f"{original_subject}"
        reply['To'] = original_to  # Respondemos al remitente
        reply['Cc'] = original_cc
       
        reply['From'] = "dickainterfaces@gmail.com"
        msg_compartido = f"Saludos,\n\nSe comparte la confirmación de recibido para el  envío entrante identificado como:{trknum}\n\nFavor de revisar el excel adjunto con los detalles de cada línea recibida."
        reply.set_content(f"{msg_compartido} \n\nMuchas Gracias.\n\nBuen día.")

        # Ruta de tu archivo Excel
        nombre_archivo = f"{folder_excel}\{trknum_limpio}.xlsx" 
        archivo        = f"Confirmación {trknum_limpio}.xlsx" 
        # 1. Leer el archivo en binario
        with open(nombre_archivo, 'rb') as f:
           file_data = f.read()
        # Para Excel suele ser: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
        maintype, subtype = "application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        reply.add_attachment(file_data, maintype=maintype, subtype=subtype,filename=archivo)  
	    
        #Enviar la respuesta
        try:
          with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
             smtp.login(username,password)
             smtp.send_message(reply)             
             print(f"Respuesta enviada a {original_to }")
        except Exception as e:
             print(f"Error al responder: {e}")  
        
        ###asincrono###
        
        
        resultado = {"Confirmación recibida,MASTER_RCPT_COMPLETE_OUB_IFD, muchas gracias.": trknum }
        return JSONResponse(
           status_code=200   , 
           content=resultado
            )
    
        #return {access_token }
    except Exception as e:
         return {"error": "Formato inválido", "detalle": str(e)}, 400
     
    
                       
if __name__ == "__main__":
   import uvicorn
   port= int(os.getenv("PORT",8000))
   uvicorn.run(app, host="0.0.0.0", port=port)
